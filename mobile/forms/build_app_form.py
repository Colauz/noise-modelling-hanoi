#!/usr/bin/env python3
"""Derive the app's XLSForm from the campaign's v2, adding the fields the app records.

The mobile app measures a level with the phone's own microphone. That is a
different quantity from `noise_db`, which is a trimmed Decibel X reading from one
of three cross-calibrated handsets (docs/metrology.md), and it must not land in
the same column. This script produces `hanoi_noise_app_v3.xlsx`, identical to v2
plus five app-filled fields, so the two quantities stay apart in Kobo as well.

Same `id_string` as v2 on purpose: Kobo treats it as a new *version* of the same
form, so the existing project keeps its submissions instead of splitting in two.

    python3 mobile/forms/build_app_form.py

Do not hand-edit the output. Change v2 or change this script.
"""
from __future__ import annotations

import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SOURCE = os.path.join(ROOT, "data", "forms", "hanoi_noise_form_v2.xlsx")
OUTPUT = os.path.join(HERE, "hanoi_noise_app_v3.xlsx")

#: Bumped whenever this script changes what it emits. Must equal
#: `NOISE_FORM_V3_VERSION` in app/src/main/java/.../form/FormSpec.kt — Kobo
#: validates the version an instance declares.
VERSION = "2026082002"

#: Appended to the `collector` choice list. The deployed v2 list is three first
#: names, which is the right list for a team of three and the wrong one for a
#: public campaign: it asks a stranger to file themselves under someone else's
#: name. `public` is the honest bucket, and it keeps the three named collectors
#: distinguishable — `01_prepare_field_data.py` keys its per-collector
#: calibration offset and its de-duplication on this field.
COLLECTOR_CHOICES = [("collector", "public", "Public contributor (mobile app)")]

# (type, name, label, hint). Order is the order they appear in the form.
APP_FIELDS = [
    ("decimal", "app_noise_db",
     "Level measured by the app (dB)",
     "Filled by the app. Uncalibrated in absolute terms — see docs/metrology.md"),
    ("text", "measure_method",
     "How the level was obtained",
     "in_app_<audio source> when the app measured it, external_meter_app otherwise"),
    ("text", "device_model", "Device", "Manufacturer and model, filled by the app"),
    ("text", "os_version", "Operating system", "Filled by the app"),
    ("text", "app_version", "App version", "Filled by the app"),
    ("text", "contributor_id",
     "Contributor",
     "A random identifier made on this phone at first launch. Not a device id, not a name"),
]


def main() -> None:
    if not os.path.isfile(SOURCE):
        sys.exit(f"missing {SOURCE}")

    shutil.copyfile(SOURCE, OUTPUT)
    book = openpyxl.load_workbook(OUTPUT)
    survey = book["survey"]

    header = [cell.value for cell in survey[1]]
    columns = {name: i for i, name in enumerate(header) if name}
    for required in ("type", "name", "label"):
        if required not in columns:
            sys.exit(f"the survey sheet has no {required!r} column")

    existing = {survey.cell(row=r, column=columns["name"] + 1).value
                for r in range(2, survey.max_row + 1)}

    added = 0
    for kind, name, label, hint in APP_FIELDS:
        if name in existing:
            continue
        row = survey.max_row + 1
        survey.cell(row=row, column=columns["type"] + 1, value=kind)
        survey.cell(row=row, column=columns["name"] + 1, value=name)
        survey.cell(row=row, column=columns["label"] + 1, value=label)
        if "hint" in columns:
            survey.cell(row=row, column=columns["hint"] + 1, value=hint)
        added += 1

    choices = book["choices"]
    choice_header = [cell.value for cell in choices[1]]
    choice_columns = {name: i for i, name in enumerate(choice_header) if name}
    for required in ("list_name", "name", "label"):
        if required not in choice_columns:
            sys.exit(f"the choices sheet has no {required!r} column")

    existing_choices = {
        (choices.cell(row=r, column=choice_columns["list_name"] + 1).value,
         choices.cell(row=r, column=choice_columns["name"] + 1).value)
        for r in range(2, choices.max_row + 1)
    }
    for list_name, name, label in COLLECTOR_CHOICES:
        if (list_name, name) in existing_choices:
            continue
        row = choices.max_row + 1
        choices.cell(row=row, column=choice_columns["list_name"] + 1, value=list_name)
        choices.cell(row=row, column=choice_columns["name"] + 1, value=name)
        choices.cell(row=row, column=choice_columns["label"] + 1, value=label)
        added += 1

    # The clip is required of the three campaign handsets and optional for
    # everyone else; the form cannot express "required for some", so it is
    # relaxed here and the app enforces it for the team.
    required_at = columns.get("required")
    if required_at is not None:
        for r in range(2, survey.max_row + 1):
            if survey.cell(row=r, column=columns["name"] + 1).value == "audio_sample":
                survey.cell(row=r, column=required_at + 1, value=0)
                added += 1

    settings = book["settings"]
    settings_header = [cell.value for cell in settings[1]]
    if "version" not in settings_header:
        sys.exit("the settings sheet has no 'version' column")
    settings.cell(row=2, column=settings_header.index("version") + 1, value=VERSION)

    book.save(OUTPUT)
    print(f"wrote {os.path.relpath(OUTPUT, ROOT)}: {added} field(s) added, version {VERSION}")


if __name__ == "__main__":
    main()
